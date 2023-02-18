import openpyxl
import json
wb=openpyxl.load_workbook('./resume.xlsx')
print(wb.get_sheet_names())
응답시트=wb['설문지 응답 시트1']
print("몇번부터 끝까지 등록할지 입력해주세요.")
전체세로=응답시트.max_column

#입력받아야함
몇번부터추가=int(input())+1

지원자들=list()
for row in range(몇번부터추가,전체세로+1):
    지원자=dict()
    for 공통 in range(2,12):
        #공통 객체 생성
        지원자[응답시트.cell(row=1, column=공통).value]=응답시트.cell(row=row, column=공통).value

    if(응답시트.cell(row=row, column=2).value=="기획"):

        for 기획 in range(12,19):
            지원자[응답시트.cell(row=1, column=기획).value]=응답시트.cell(row=row, column=기획).value

        지원자들.append(지원자)

    if(응답시트.cell(row=row, column=2).value=="디자인"):

        for 디자인 in range(19,27):
            지원자[응답시트.cell(row=1, column=디자인).value]=응답시트.cell(row=row, column=디자인).value

        지원자들.append(지원자)

    if(응답시트.cell(row=row, column=2).value=="프론트엔드"):

        for 프론트엔드 in range(27,35):
            지원자[응답시트.cell(row=1, column=프론트엔드).value]=응답시트.cell(row=row, column=프론트엔드).value

        지원자들.append(지원자)

    if(응답시트.cell(row=row, column=2).value=="백엔드"):

        for 백엔드 in range(35,43):
            지원자[응답시트.cell(row=1, column=백엔드).value]=응답시트.cell(row=row, column=백엔드).value

        지원자들.append(지원자)
    
with open('./지원자.json','w',encoding='utf-8') as f:
    json.dump(지원자들, f, ensure_ascii=False)